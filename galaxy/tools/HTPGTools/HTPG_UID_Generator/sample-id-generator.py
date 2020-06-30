#!/usr/bin/env python

#Description:
#This script generates a give number of UIDs which are then added to the HTPG Sample file template file and the Sample list in the Intertek Order form.

#For more info: wkigoni@gmail.com

#Quick usage:
#python sample-id-generator.py --samplefile=/path/to/samplefile.xlsx --orderfile=/path/to/orderfile.xlsx --samplecount=50


#Full usage:
#usage: sample-id-generator.py [-h] [--samplefile SAMPLEFILE]
#                              [--orderfile ORDERFILE]
#                              [--samplecount SAMPLECOUNT]
#                              [--sampleidchars SAMPLEIDCHARS] [--debug DEBUG]
#                              [--samplefile_sampleid_start_cell SAMPLEFILE_SAMPLEID_START_CELL]
#                              [--samplefile_sampleid_worksheet SAMPLEFILE_SAMPLEID_WORKSHEET]
#                              [--orderfile_sampleid_start_cell ORDERFILE_SAMPLEID_START_CELL]
#                              [--orderfile_sampleid_worksheet ORDERFILE_SAMPLEID_WORKSHEET

import os
import argparse
import hashlib
import logging

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


DEF_SAMPLE_COUNT = 5
DEF_SAMPLEID_MAX_CHARS = 12
DEF_ORDERFILE = "Intertek_Order_Form.xlsx"
DEF_SAMPLEFILE = "HTPG_Sample_File.xlsx"
DEF_SAMPLE_FILE_SAMPLEID_START_CELL = 'A2'
DEF_SAMPLEFILE_SAMPLEID_WORKSHEET_NAME = 'Sample_file'
DEF_ORDERFILE_SAMPLEID_START_CELL = 'B16'
DEF_ORDERFILE_SAMPLEID_WORKSHEET_NAME = 'Sample List'

DEBUG = True

class SampleIDGenerator(object):

    def __init__(self,
                 samplefile,
                 orderfile,
                 samplecount,
                 debug,
                 sampleidchars,
                 samplefile_sampleid_start_cell,
                 samplefile_sampleid_worksheet,
                 orderfile_sampleid_start_cell,
                 orderfile_sampleid_worksheet,
                 **kwargs
                 ):
        self.samplefile = samplefile
        self.orderfile = orderfile
        self.samplecount = int(samplecount)
        self.debug = debug
        self.sampleidchars = sampleidchars
        self.samplefile_sampleid_start_cell = samplefile_sampleid_start_cell
        self.samplefile_sampleid_worksheet = samplefile_sampleid_worksheet
        self.orderfile_sampleid_start_cell = orderfile_sampleid_start_cell
        self.orderfile_sampleid_worksheet = orderfile_sampleid_worksheet

        for key, value in kwargs.items():
            setattr(self, key, value)
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
 
    """
        Generate x no of samples
    """
    def generate_sampleIDs(self, samplescount):
        sampleIDs=list()
        for i in range(samplescount):
            sampleID = hashlib.sha1(os.urandom(128)).hexdigest()[:self.sampleidchars]
            if len(sampleID) < self.sampleidchars:
                log_str = "sampleID: {sampleID} less than minimum chars of {chars}".format(
                        sampleID=sampleID, chars=self.sampleidchars
                    )
                if self.debug: self.logger.info(log_str)
                pass
            else:
                if self.debug: print "generated sampleID: {ID}".format(ID=sampleID)
                sampleIDs.append(sampleID)
        return sampleIDs

    """
        Add generated sampleIDs in the sample file
    """
    def save_sampleIDs(self, samplefile, sampleIDs,):
        start_cel=self.samplefile_sampleid_start_cell
        sampleIDs_worksheet_name=self.samplefile_sampleid_worksheet
        workbook = load_workbook(filename=samplefile)
        sampleID_ws = workbook[sampleIDs_worksheet_name]
        
        start_col = start_cel[0]
        start_row = int(start_cel[1])
        end_row = start_row + len(sampleIDs)
        counter = 0

        for row in range(start_row, end_row):
            cell = "{col}{row}".format(col=start_col, row=row)
            log_str = "{col}{row}={sampleID}".format(col=start_col, row=row, sampleID=sampleIDs[counter])
            if self.debug: self.logger.info(log_str)
            sampleID_ws[cell] = sampleIDs[counter]
            counter += 1
        workbook.save(filename = samplefile)
        if self.debug: self.logger.info("successfully saved sample IDs in {}".format(samplefile))

    def save_orderfile(self, sampleIDs, orderfile,):
        start_cel_orderfile=self.orderfile_sampleid_start_cell
        sampleIDs_worksheet_name=self.orderfile_sampleid_worksheet
        workbook = load_workbook(filename=orderfile)
        sampleID_ws = workbook[sampleIDs_worksheet_name]

        start_col = start_cel_orderfile[0]
        start_row = int(start_cel_orderfile[1:])
        end_row = start_row + len(sampleIDs)
        counter = 0

        for row in range(start_row, end_row):
            cell = "{col}{row}".format(col=start_col, row=row)
            log_str = "{col}{row}={sampleID}".format(col=start_col, row=row, sampleID=sampleIDs[counter])
            if self.debug: self.logger.info(log_str)
            sampleID_ws[cell] = sampleIDs[counter]
            counter += 1
        workbook.save(filename=orderfile)
        if self.debug: self.logger.info("successfully saved sample IDs in {}".format(orderfile))

    def do_stuff(self):
        sampleIDs = self.generate_sampleIDs(self.samplecount)
        self.save_sampleIDs(samplefile=self.samplefile, sampleIDs=sampleIDs)
        self.save_orderfile(orderfile=self.orderfile, sampleIDs=sampleIDs)


def command():
    parser = argparse.ArgumentParser(description='SampleID generator script')
    parser.add_argument(
        '--samplefile',
        type=str,
        help='path to sample file',
		default=DEF_SAMPLEFILE,
        dest='samplefile')
    parser.add_argument(
        '--orderfile',
        help='path to order file',
		default=DEF_ORDERFILE,
        dest='orderfile')
    parser.add_argument(
        '--samplecount',
        help='number of sample IDs to generate',
        default=DEF_SAMPLE_COUNT,
        dest='samplecount')
    parser.add_argument(
        '--sampleidchars',
        help='Number/length of sampleID characters',
        default=DEF_SAMPLEID_MAX_CHARS,
        dest='sampleidchars')
    parser.add_argument(
        '--debug',
        help='Enable/Disable Debug',
        default=DEBUG,
        dest='debug')
    parser.add_argument(
        '--samplefile_sampleid_start_cell',
        help='Samplefile SampleID start cell',
        default=DEF_SAMPLE_FILE_SAMPLEID_START_CELL,
        dest='samplefile_sampleid_start_cell')
    parser.add_argument(
        '--samplefile_sampleid_worksheet',
        help='Samplefile SampleID worksheet name',
        default=DEF_SAMPLEFILE_SAMPLEID_WORKSHEET_NAME,
        dest='samplefile_sampleid_worksheet')
    parser.add_argument(
        '--orderfile_sampleid_start_cell',
        help='Orderfile SampleID start cell',
        default=DEF_ORDERFILE_SAMPLEID_START_CELL,
        dest='orderfile_sampleid_start_cell')
    parser.add_argument(
        '--orderfile_sampleid_worksheet',
        help='Orderfile SampleID worksheet name',
        default=DEF_ORDERFILE_SAMPLEID_WORKSHEET_NAME,
        dest='orderfile_sampleid_worksheet')

    args, unknown = parser.parse_known_args()
    parms_to_return = {}

    for arguments in dir(args):
        if not arguments.startswith("_"):
            if getattr(args, arguments) is None:
                raise Exception("{} is required".format(arguments))
            else:
                parms_to_return[arguments] = getattr(args, arguments)

    for i, j in enumerate(unknown):
        if j.startswith("--"):
            param_name = j.replace("--", "", 1)
            parms_to_return[param_name] = unknown[i + 1]

    return parms_to_return

def main():
    SampleIDGenerator(**command()).do_stuff()


if __name__ == '__main__':
    main()
